import json
import pandas as pd
from app.core.utils.Logger import Logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

logger = Logger(name="FileMaker", log_file="logs/utils/FileMaker.log").get_logger()


class FileMaker:
    def __init__():

        pass

    @staticmethod
    def save_list_to_json(list):
        if len(list) == 0:
            logger.info("json파일저장실패, 저장된 링크 없음")
            return False

        json_file_path = ".data/musinsa_event_links.json"

        with open(json_file_path, "w", encoding="utf-8") as json_file:
            json.dump(list, json_file, ensure_ascii=False, indent=4)

        print(f"총 {len(list)}개의 링크를 {json_file_path} 파일에 저장완료")

    @staticmethod
    def save_to_excel_for_musinsa(
        infos_list, file_name="infos_list", column_order=None
    ):
        """

        - params

        column_order: str[] e.g. [
            "브랜드",
            "영문명",
            "키프리스 바로가기" "상호 / 대표자",
            "브랜드 페이지",
            "연락처",
            "E-mail",
            "사업자번호",
            "통신판매업신고",
            "영업소재지",
        ]

        """
        # DataFrame 생성
        df = pd.DataFrame(infos_list)

        # 칼럼 순서 지정
        try:

            if column_order:
                df = df[column_order]
            else:
                df = df[df.columns]

        except Exception as e:
            logger.exception('칼럼오더를 데이터 프레임 변환 시 오류, {e}')
            
        # 엑셀 파일로 임시 저장
        temp_file = f"{file_name}_temp.xlsx"
        df.to_excel(temp_file, index=False)

        # openpyxl을 사용하여 하이퍼링크 추가
        workbook = load_workbook(temp_file)
        sheet = workbook.active

        # "브랜드 페이지" 열 위치 찾기
        brand_page_column = None
        for col in sheet.iter_cols(1, sheet.max_column, 1, 1):
            if col[0].value == "브랜드 페이지":
                brand_page_column = col[0].column_letter
                break

        if brand_page_column is None:
            raise ValueError("'브랜드 페이지'라는 열을 찾을 수 없습니다.")

        # "브랜드 페이지" 열에서 URL을 하이퍼링크로 설정
        for row in range(2, sheet.max_row + 1):  # 헤더 이후 데이터 행만 처리
            cell = sheet[f"{brand_page_column}{row}"]
            url = cell.value
            if url:  # URL 값이 존재하는 경우
                cell.value = "바로가기"  # 셀 텍스트 설정
                cell.hyperlink = url  # 하이퍼링크 설정
                cell.font = Font(
                    color="0000FF", underline="single"
                )  # 하이퍼링크 스타일 적용

        # 최종 엑셀 파일 저장
        workbook.save(f"{file_name}.xlsx")
        workbook.close()